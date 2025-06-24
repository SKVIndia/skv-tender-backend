from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer, util
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

app = Flask(__name__)
CORS(app)

@app.route("/api/compare", methods=["POST"])
def compare_files():
    if "skv_file" not in request.files or "tender_file" not in request.files:
        return jsonify({"error": "Both files are required"}), 400

    try:
        skv_file = request.files["skv_file"]
        tender_file = request.files["tender_file"]

        skv_df = pd.read_excel(skv_file)
        tender_df = pd.read_excel(tender_file)

        skv_clauses = skv_df[['Clauses', 'SKV Standard']].dropna()
        tender_brief = tender_df.iloc[1:, [1, 2, 3]]
        tender_brief.columns = ['Tender Brief', 'Value', 'Doc Name and Page Number']
        tender_brief = tender_brief.dropna()

        model = SentenceTransformer('all-MiniLM-L6-v2')
        skv_embeddings = model.encode(skv_clauses['Clauses'].tolist(), convert_to_tensor=True)
        tender_embeddings = model.encode(tender_brief['Tender Brief'].tolist(), convert_to_tensor=True)

        results = []
        matched_tender_indices = set()

        for i, skv_clause in skv_clauses.iterrows():
            cosine_scores = util.cos_sim(skv_embeddings[i], tender_embeddings)[0]
            best_match_idx = int(np.argmax(cosine_scores))
            score = float(cosine_scores[best_match_idx])
            tender_row = tender_brief.iloc[best_match_idx]
            matched_tender_indices.add(best_match_idx)

            if score > 0.85:
                inference = "✅ Match"
                fill_color = "C6EFCE"
            elif 0.6 < score <= 0.85:
                inference = "🟡 Needs Clarification"
                fill_color = "FFF2CC"
            else:
                inference = "❌ Conflict or Not Found"
                fill_color = "F4CCCC"

            results.append({
                "SKV Standards": f"{skv_clause['Clauses']}: {skv_clause['SKV Standard']}",
                "Tender Brief": f"{tender_row['Tender Brief']}: {tender_row['Value']}",
                "Inference": inference,
                "Doc Name and Page Number": tender_row['Doc Name and Page Number'],
                "Fill Color": fill_color
            })

        comparison_df = pd.DataFrame(results)
        fill_colors = comparison_df.pop("Fill Color")

        extra_rows = []
        for i, row in tender_brief.iterrows():
            if i not in matched_tender_indices:
                extra_rows.append({
                    "Tender Brief Extra Field": row['Tender Brief'],
                    "Value": row['Value'],
                    "Doc Name and Page Number": row['Doc Name and Page Number'],
                    "Comment": "Not part of SKV Standards"
                })
        extra_df = pd.DataFrame(extra_rows)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, index=False, sheet_name="SKV vs Tender")
            extra_df.to_excel(writer, index=False, sheet_name="Extra Tender Fields")

            wb = writer.book
            ws = writer.sheets["SKV vs Tender"]
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3)):
                color = fill_colors.iloc[i]
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for cell in row:
                    cell.fill = fill
                    cell.font = Font(color="000000")

            ws_extra = writer.sheets["Extra Tender Fields"]
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            for row in ws_extra.iter_rows(min_row=2, max_row=ws_extra.max_row, min_col=1, max_col=4):
                for cell in row:
                    cell.fill = yellow_fill
                    cell.font = Font(color="000000")

        output.seek(0)
        return send_file(output, as_attachment=True,
                         download_name="SKV_Tender_Comparison_Result.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        return jsonify({"error": f"Processing failed: {str(e)}"}), 500

@app.route("/", methods=["GET"])
def health_check():
    return jsonify({"status": "Backend is running"}), 200

if __name__ == "__main__":
    app.run(debug=True, port=5000)
